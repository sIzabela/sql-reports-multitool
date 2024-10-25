import os
import logging
import shutil
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import create_engine
from openpyxl import load_workbook
from dotenv import load_dotenv

# base_path = "C://Users//robot//Desktop"
base_path = "C://Users//izabe//OneDrive//Pulpit"
def setup_logging():
    # Uzyskanie dzisiejszej daty w formacie yyyyMMdd
    today_date = datetime.now().strftime("%Y%m%d")
    log_file_path = f"./logs/{today_date}.log"
    
    # Sprawdzanie, czy plik istnieje
    if not os.path.exists(log_file_path):
        # Jeśli plik nie istnieje, tworzenie nowego pliku
        os.makedirs(os.path.dirname(log_file_path), exist_ok=True)
        with open(log_file_path, 'w') as file:
            pass # Tworzenie pustego pliku
    
    # Ustawianie zapisywania logów w tym pliku
    import logging
    logging.basicConfig(filename=log_file_path, level=logging.INFO, format='%(asctime)s - %(message)s', encoding='utf-8')

def log_message(message):
    logging.info(message)
    print("\nLog: " + message)

def remove_old_logs():

    log_message(f"Usuwam pliki starsze niż 30 dni")
    logs_dir = "./logs"
    # Definiowanie progu czasu dla plików starszych niż 30 dni
    threshold_date = datetime.now() - timedelta(days=30)
    threshold_time = threshold_date.timestamp()
    
    # Przechodzenie przez pliki w katalogu
    for filename in os.listdir(logs_dir):
        file_path = os.path.join(logs_dir, filename)
        
        if os.path.isfile(file_path):
            # Sprawdzanie czasu utworzenia pliku
            file_create_time = os.path.getctime(file_path)
            
            if file_create_time < threshold_time:
                # Usuwanie plików starszych niż 30 dni
                os.remove(file_path)
                log_message(f"Usunięto: {file_path}")

def input_valid_date():
    while True:
        try:
            date = input("Podaj datę (RRRRMMDD) lub nie wpisuj nic, aby zakończyć: ")
            if not date:
                return None
            else:
                datetime.strptime(date, '%Y%m%d')
                return date
        except ValueError:
            print("Błędny format daty. Spróbuj ponownie")

def pobieranie_eksportowanie_danych(query, engine, output_path):
    # Pobieranie danych z SQL
    try:
        data = pd.read_sql(query, engine)
        log_message(">>>> Dane pobrane")
    except Exception as e:
        log_message(f">>>> Błąd podczas pobierania danych: {e}")
        return

    # Eksportowanie danych do pliku Excel
    try:
        data.to_excel(output_path, index=False)
        log_message(f">>>> Dane wyeksportowane do pliku Excel: {output_path}")
    except Exception as e:
        log_message(f">>>> Błąd podczas eksportowania danych do pliku Excel: {e}")
        return

def SQL(view_name, output_path):
    server_name = os.getenv('server_name')
    user = os.getenv('user')
    passwd = os.getenv('passwd')
    database_name = os.getenv('database_name')

    # Łączenie z bazą danych
    try:
        connection_string = f"mssql+pyodbc://{user}:{passwd}@{server_name}/{database_name}?driver=SQL+Server&TrustServerCertificate=Yes"
        engine = create_engine(connection_string)
        query = f"SELECT * FROM {view_name}"
        log_message(">>> Połączono z bazą danych")
    except Exception as e:
        log_message(f">>> Błąd podczas łączenia z bazą danych: {e}")
        return

    pobieranie_eksportowanie_danych(query, engine, output_path)

def excel_naglowki(output_path, headers, startColumn):
        # Sprawdzenie, czy plik Excel został poprawnie zapisany
    if os.path.exists(output_path):
        # Dodawanie nagłówków w raporcie
        try:
            workbook = load_workbook(output_path)
            worksheet = workbook.active
            
            for index, header in enumerate(headers, start=startColumn):
                worksheet.cell(row=1, column=index).value = header

            workbook.save(output_path)
            log_message(f">>> Nagłówki dodane do kolumn od {startColumn}. w pliku Excel")
        except Exception as e:
            log_message(f">>> Błąd podczas dodawania nagłówków do kolumn w pliku Excel: {e}")
    else:
        log_message(f">>> Plik Excel nie został znaleziony: {output_path}")

def ryzyka():
    view_name = os.getenv('view_name_ryzyka')

    date = datetime.now().strftime("%Y%m%d")
    output_path = f"{base_path}//AC_ryzyka//{date}//RaportAcRyzyka_{date}.xlsx"

    if os.path.exists(output_path):
        log_message(f">>> Plik {output_path} już istnieje.")
    else:
        SQL(view_name, output_path)
        headers = ['Finish_type','Rejection_reason','Process_repeat','Send_mail','Send_sms','Sms_delivered','caseID','Status_robot']
        startColumn = 28
        excel_naglowki(output_path, headers, startColumn)

def historia_pobieranie_dzisiaj():
    view_name = os.getenv('view_name_historia')

    date = datetime.now().strftime("%Y%m%d")
    output_path = f"{base_path}//daneABP//{date}//RaportDaneABP_{date}.xlsx"

    if os.path.exists(output_path):
        log_message(f">>> Plik {output_path} już istnieje.")
    else:
        SQL(view_name, output_path)
        headers = ['gov_PROGNOZOWANY_PRZEBIEG','gov_POCHODZENIE_POJAZDU','gov_WYKORZYSTANIE_POJAZDU','gov_BADANIE_TECHNICZNE','gov_ZMIANY_WLASCICIELA','gov_WSPOLWLASCICIEL','info_PROGNOZOWANY_PRZEBIEG','info_POCHODZENIE_POJAZDU','info_WYKORZYSTANIE_POJAZDU','info_BADANIE_TECHNICZNE','info_ZMIANY_WLASCICIELA','info_WSPOLWLASCICIEL','PODSUMOWANIE','UWAGI','DATA_RAPORTU']
        startColumn = 15
        excel_naglowki(output_path, headers, startColumn)         

def historia_pobieranie_data():
    server_name = os.getenv('server_name')
    user = os.getenv('user')
    passwd = os.getenv('passwd')
    database_name = os.getenv('database_name')
    sql_query = os.getenv('sql')

    date = input_valid_date()
    if date == None:
        return

    output_path = f"{base_path}//daneABP//{date}//RaportDaneABP_{date}.xlsx"

    if os.path.exists(output_path):
        log_message(f">>> Plik {output_path} już istnieje.")
    else:
        log_message(f">>> Pobieranie raportu z dnia {date}")

        raport_date = (datetime.strptime(date, '%Y%m%d') - timedelta(days=1)).strftime('%Y%m%d')
        today = datetime.now()
        delta = today - datetime.strptime(raport_date, '%Y%m%d')

        # Łączenie z bazą danych
        try:
            connection_string = f"mssql+pyodbc://{user}:{passwd}@{server_name}/{database_name}?driver=SQL+Server&TrustServerCertificate=Yes"
            engine = create_engine(connection_string)
            query =(sql_query.format(delta=delta)
            )
            log_message(">>> Połączono z bazą danych")
        except Exception as e:
            log_message(f">>> Błąd podczas łączenia z bazą danych: {e}")
            return
        
        pobieranie_eksportowanie_danych(query, engine, output_path)

        headers = ['gov_PROGNOZOWANY_PRZEBIEG','gov_POCHODZENIE_POJAZDU','gov_WYKORZYSTANIE_POJAZDU','gov_BADANIE_TECHNICZNE','gov_ZMIANY_WLASCICIELA','gov_WSPOLWLASCICIEL','info_PROGNOZOWANY_PRZEBIEG','info_POCHODZENIE_POJAZDU','info_WYKORZYSTANIE_POJAZDU','info_BADANIE_TECHNICZNE','info_ZMIANY_WLASCICIELA','info_WSPOLWLASCICIEL','PODSUMOWANIE','UWAGI','DATA_RAPORTU']
        startColumn = 15
        excel_naglowki(output_path, headers, startColumn)
       
def historia_scalanie():
    print("\nMenu wyboru Scalanie - Historia Pojazdu:")
    print("1 - Scalanie raportów PO WEEKENDZIE (z trzech dni)")
    print("2 - Scalanie wskazanych raportów")
    print("3 - Powrót do poprzedniego menu")
    choice_scalanie = input("Wybierz opcję (1/2/3): ")

    if choice_scalanie == "1":
        log_message(">>> Wybrano opcję: Scalanie PO WEEKENDZIE")

        # Dziś
        today = datetime.now().strftime("%Y%m%d")
        today_folder = f"{base_path}//daneABP//{today}"
        today_file = f"RaportDaneABP_{today}.xlsx"
        today_path = f"{today_folder}//{today_file}"

        # Wczoraj
        yesterday = datetime.now().strftime("%Y%m%d") - timedelta(days=1)
        yesterday_path = f"{base_path}//daneABP//{yesterday}//RaportDaneABP_{yesterday}.xlsx"

        # Przedwczoraj
        two_days_ago = datetime.now().strftime("%Y%m%d") - timedelta(days=2)
        two_days_ago_path = f"{base_path}//daneABP//{two_days_ago}//RaportDaneABP_{two_days_ago}.xlsx"

        weekend = [today_path, yesterday_path, two_days_ago_path]
        pathList = []
        merge = None
        for day in weekend:
            pathList.append(day)
            filesData = pd.read_excel(day)
            merge = pd.concat([merge, filesData])

        merge.to_excel(today_path, index=False)

        for path in pathList:
            folder_to_delete = os.path.dirname(path)
            shutil.move(path, os.path.join(today_folder, os.path.basename(path).replace('.xlsx', f'_scalony.xlsx')))
            if os.path.exists(folder_to_delete):
                shutil.rmtree(folder_to_delete)

        log_message(f">>>> Scalone pliki zostały przeniesione do folderu {today_folder},a scalony plik został nazwany {today_file}.")

    elif choice_scalanie == "2":
        today = datetime.now().strftime("%Y%m%d")
        today_folder = f"{base_path}//daneABP//{today}"
        today_file = f"RaportDaneABP_{today}.xlsx"
        today_path = f"{today_folder}//{today_file}"

        log_message(">>> Wybrano opcję: Scalanie wskazanych raportów")
        print("UWAGA! Jeśli nie scalasz dzisiejszego raportu, to zabezpiecz go przed nadpisaniem scalonymi danymi")

        today = datetime.now().strftime("%Y%m%d")
        today_folder = f"{base_path}//daneABP//{today}"
        today_file = f"RaportDaneABP_{today}.xlsx"
        today_path = f"{today_folder}//{today_file}"
        
        fileCount = input("Ile plików chcesz scalić? ")
        log_message(">>>> Podano ilość plików do scalenia: " + fileCount)

        pathList = []
        merge = None
        for i in range(int(fileCount)):
            input_file = input(f"przeciągnij plik {i}: ")
            log_message(f">>> Plik {i}: {input_file}")
            pathList.append(input_file)
            filesData = pd.read_excel(input_file)
            merge = pd.concat([merge, filesData])

        for path in pathList:
            folder_to_delete = os.path.dirname(path)
            new_path = os.path.join(today_folder, os.path.basename(path).replace('.xlsx', '_scalony.xlsx'))
            shutil.move(path, new_path)
            # Remove the folder only if it is empty
            if os.path.exists(folder_to_delete) and not os.listdir(folder_to_delete):
                shutil.rmtree(folder_to_delete)
        
        merge.to_excel(today_path, index=False)

        log_message(f">>>> Scalone pliki zostały przeniesione do folderu {today_folder},a scalony plik został nazwany {today_file}.")

    elif choice_scalanie == "3":
        return
    
    else:
        print(">>> Nieprawidłowa opcja, spróbuj ponownie.")
        
def display_menu():
    print("\nMenu wyboru:")
    print("1 - Ryzyka - pobieranie raportu")
    print("2 - Historia Pojazdu")
    print("3 - Zakończ działanie")

def main():
    os.system('cls' if os.name == 'nt' else 'clear')
    load_dotenv()
    setup_logging()
    log_message("Uruchomienie aplikacji")
    
    while True:
        display_menu()
        choice = input("Wybierz opcję (1/2/3): ")

        if choice == '1':
            os.system('cls' if os.name == 'nt' else 'clear')
            log_message("> Wybrano opcję: Ryzyka - pobieranie raportu")
            try:
                ryzyka()
            except Exception as e:
                os.system('cls' if os.name == 'nt' else 'clear')
                log_message(f"> Bład wyboru opcji Ryzyka - pobieranie raportu: {e}")

        elif choice == '2':
            log_message("> Wybrano opcję: Historia Pojazdu")
            print("\nMenu wyboru Historia Pojazdu:")
            print("1 - Pobieranie dzisiejszego raportu")
            print("2 - Pobieranie raportu ze wskazanej daty")
            print("3 - Scalanie raportów")
            print("4 - Powrót do poprzedniego menu")
            subchoice = input("Wybierz opcję (1/2/3/4): ")

            if subchoice == '1':
                log_message(">> Wybrano opcję: Pobieranie dzisiejszego raportu")
                try:
                    historia_pobieranie_dzisiaj()
                except Exception as e:
                    log_message(f">> Bład opcji Pobieranie dzisiejszego raportu: {e}")

            elif subchoice == '2':
                log_message(">> Wybrano opcję: Pobieranie raportu ze wskazanej daty")
                try:
                    historia_pobieranie_data()
                except Exception as e:
                    log_message(f">> Bład opcji Pobieranie raportu ze wskazanej daty: {e}")

            elif subchoice == '3':
                log_message(">> Wybrano opcję: Scalanie raportów")
                try:
                    historia_scalanie()
                except Exception as e:
                    log_message(f">> Bład opcji Scalanie raportów: {e}")

            elif subchoice == '4':
                break

            else:
                print("> Nieprawidłowa opcja, spróbuj ponownie.")

        elif choice == '3':
            log_message("> Wybrano opcję: Zakończ działanie")
            remove_old_logs()
            break

        else:
            print("> Nieprawidłowa opcja, spróbuj ponownie.")

if __name__ == "__main__":
    main()